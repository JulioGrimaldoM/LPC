import requests, time, re
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook

def Datos_Daniel(pagina):

    try:
        libro= Workbook()
        pa=libro.active
        pa.title="Informacion"
        pa["A1"]="Nombre"
        pa["B1"]="Fecha"
        pa["C1"]="Cuidad"
        pa["D1"]="Estado"
        print("Se logro")
        libro.active=0
    except:
        print("No se pudo crear")
    count=2
    Datos={
        "Nombre":"",
        "Fecha":"",
        "Cuidad":"",
        "Estado":"",
        }
    Nombre=re.compile(r"(\w+ Corral)")
    lugar_Nac=re.compile(r"(nació\s)(en\s)(\w+)")
    fecha=re.compile(r"(\d{1,2}\s\w{1,2}\senero\s\w+\s\d{4})")
    Estado=re.compile(r"(Baja\s\w+)")
    url=pagina
    pagina=requests.get(url)
    if pagina.status_code==200:
        print("Pagina encontrada")
        soup=bs(pagina.content,"html.parser")
        info=soup.find_all("p")
        i=0
        for title in info:
            po=info[i].getText()
            lugar=lugar_Nac.search(po)
            fe=fecha.search(po)
            Est=Estado.search(po)
            Nom=Nombre.search(po)
            if Nom!=None:
                Datos["Nombre"]=Nom.group()
                pa.cell(count,1,Nom.group())
            if fe!=None:
                Datos["Fecha"]=fe.group()
                pa.cell(count,2,fe.group())
            if lugar!=None:
                Datos["Cuidad"]=lugar.group(3)
                pa.cell(count,3,lugar.group(3))
            if Est!=None:
                Datos["Estado"]=Est.group()
                pa.cell(count,4,Est.group())
            i=i+1
        print(Datos)
        i=1
        for x, y in Datos.items():
            if y=="":
                dec=input("Quieres agregar algo en "+ x+": y/n ")
                if dec!="n":
                    var=input("Escribe el valor de "+x+" es: ")
                    Datos[x]=var
                    pa.cell(count,i,var)
                    print(Datos)
                else:
                    print(Datos)
                    break
            i+=1
        libro.save("Informacion.xlsx")
if __name__=="__main__":
    pag=input("Escribe la pagina")
    Datos_Daniel(pag)
