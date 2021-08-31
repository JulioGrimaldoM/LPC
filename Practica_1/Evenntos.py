import requests
import json
import datetime
from openpyxl import Workbook

def event():
    try:
        lib= Workbook()
        pad=lib.active
        pad.title="Noticias"
        pad.title="Eventos"
        pad["A1"]="Lugar"
        pad["B1"]="Datetime"
        pad["C1"]="Datetime"
        pad["D1"]="Temperatura"
        pad["E1"]="Fecha del evento"
        print("Se logro")
        lib.active=0
    except:
        print("No se pudo crear")
    count=2
    
#definir 3 fechas para eventos del atleta 
#monterrey 20 de mayo
##Seattle 21 de mayo
#bogota 22 de mayo
    print('\n')
    print("Tenemos que apoyar a nuestro querido Daniel Corral en sus proximos eventos")
    print("lo que sabemos de ellos es:")
    print('\n')

#para monterrey
    evento="Monterrey MX"
    print(evento)
    dia="20 de mayo del 2021"
    lat = "25.725536"
    lon = "-100.315157"
    appid = "33963f637195707ae5cd64ff8aabfa8c" 
    page = requests.get ("https://api.openweathermap.org/data/2.5/onecall?lat="
                         +lat+"&lon="+lon
                         +"&exclude=minutely,hourly,alert&units=metric&appid="+appid)

    print (page.status_code)                                

    weatherData = json.loads(page.content)

    i=0
    for dt in weatherData["daily"]:
        print(weatherData["daily"][i]["dt"]) #horario en UNIX -> 1/Enero/1970 en UTC
        pad.cell(count,1,evento)
        dt = int(weatherData["daily"][i]["dt"])
        pad.cell(count,2,dt)
        amanecer = int(weatherData["daily"][i]["sunrise"])
        pad.cell(count,3,amanecer)
        atardecer = int(weatherData["daily"][i]["sunset"])
        pad.cell(count,4,atardecer)
        i+=1
        TempEvento=weatherData["daily"][3]["temp"]
        pad.cell(count,4,str(TempEvento))
        pad.cell(count,5,dia)
        count+=1
        break
    print("La temperatura para el evento del dia 20 de mayo es de ")
    print(TempEvento)
    print('\n')

#Para Seattle [47.6062, -122.3321]

    evento1="Seattle US"
    print(evento1)

    dia1="21 de mayo del 2021"


    lat = "47.6062"
    lon = "-122.3321"
    appid = "33963f637195707ae5cd64ff8aabfa8c" 
    page = requests.get ("https://api.openweathermap.org/data/2.5/onecall?lat="
                         +lat+"&lon="+lon
                         +"&exclude=minutely,hourly,alert&units=metric&appid="+appid)
                    
    print (page.status_code)                               
    weatherData = json.loads(page.content)

    i=0
    for dt in weatherData["daily"]:
        print(weatherData["daily"][i]["dt"]) #horario en UNIX -> 1/Enero/1970 en UTC
        dt = int(weatherData["daily"][i]["dt"])   
        amanecer = int(weatherData["daily"][i]["sunrise"])      
        atardecer = int(weatherData["daily"][i]["sunset"])  
        i+=1
        TempEvento1=weatherData["daily"][4]["temp"]
        pad.cell(count,1,evento1)
        pad.cell(count,2,dt)
        pad.cell(count,3,amanecer)
        pad.cell(count,4,atardecer)
        pad.cell(count,4,str(TempEvento1))
        pad.cell(count,5,dia1)
        count+=1
        break
    print("La temperatura para el evento del dia 21 de mayo es de ")
    print(TempEvento1)
    print('\n')

#bogotá[4.6097, -74.0817]

    evento2="Bogotá CO"
    print(evento2)

    dia1="22 de mayo del 2021"
    lat = "4.6097"
    lon = "-74.0817"
    appid = "33963f637195707ae5cd64ff8aabfa8c" 
    page = requests.get ("https://api.openweathermap.org/data/2.5/onecall?lat="
                         +lat+"&lon="+lon
                         +"&exclude=minutely,hourly,alert&units=metric&appid="+appid)
                    
    print (page.status_code)                                

    weatherData = json.loads(page.content)
    i=0
    for dt in weatherData["daily"]:
        print(weatherData["daily"][i]["dt"]) #horario en UNIX -> 1/Enero/1970 en UTC
        dt = int(weatherData["daily"][i]["dt"])  
        amanecer = int(weatherData["daily"][i]["sunrise"]) 
        atardecer = int(weatherData["daily"][i]["sunset"])  
        i+=1
        TempEvento2=weatherData["daily"][5]["temp"]
        pad.cell(count,1,evento2)
        pad.cell(count,2,dt)
        pad.cell(count,3,amanecer)
        pad.cell(count,4,atardecer)
        pad.cell(count,4,str(TempEvento2))
        pad.cell(count,5,dia1)
        count+=1
        break
    print("La temperatura para el evento del dia 22 de mayo es de ")
    print(TempEvento2)
    print('\n')
    lib.save("Eventos.xlsx")
if __name__=="__main__":        
    event()
